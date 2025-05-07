# -*- coding: utf-8 -*-
"""
Created on Sat Apr 26 19:50:09 2025

@author: krati
Children's Vulnerability Scale
code to assign arbitray numerical value to different answers in the questionairre 
data is stored in an excel sheet and has an answer from a finite set of options 
We are going to assign a number to every response and save the data in a different sheet

problem
"""

import functions as fn
import openpyxl



wb = openpyxl.load_workbook('Respsheet.xlsx') #write your file name
ws = wb['vresan']#The name of the sheet where data is stored
wsout = wb['CtS']
row = ws.max_row  #number of children
column = ws.max_column  #number of questions
""" add the heading to new sheet
Working
"""
values = [ws.cell(row=1,column=i).value for i in range(1,column+1)]
for k in range(len(values)):
    wsout.cell(row = 1, column = k+1).value = values[k]
"""
Skip Column 6,8,1,2,3,10,14,27,31
this is a temp measure
skip = [1,2,3,6,8,10,14,27,31]
"""
#temp measure for bad data

demo = [1,2,3,6,8]
#loop to assign numerical values to data
for i in range(2,row+1):
         for j in range(1,column+1):
             
             if j in demo:
                     val = ws.cell(row=i,column=j).value 
                     wsout.cell(row=i, column=j, value = val)                     
                 
             else:
                # copy the data in new sheet as a number
                fn.assign_score(ws, i, j, wsout, i, j)
print("done")
wb.save('Respsheet.xlsx')
