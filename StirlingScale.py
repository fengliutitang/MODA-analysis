# -*- coding: utf-8 -*-
"""
Created on Sun Nov  3 19:21:30 2024

The Stirling Children’s Wellbeing Scale 
Higher Score indicates better wellbeing
Highest Score = 60
Lowest Score = 15
Done
"""


import openpyxl 

wb = openpyxl.load_workbook('Respsheet.xlsx') #write your file name
ws = wb['WBres_an']#The name of the sheet where data is stored
row = ws.max_row #number of children
column = ws.max_column #number of questions
print(row,column)
score = 0
pes_score = 0
po_score =0
sdss_score =0
ws.cell(row=1, column=column+2, value = 'Total score')
ws.cell(row=1, column=column+3, value = 'PES score')
ws.cell(row=1, column=column+4, value = 'PO score')
ws.cell(row=1, column=column+5, value = 'SDSS score')
skip = [3,8,14]
pes = [10,15,16,11,13,12]
po = [9,6,2,5,7,4]
for i in range(2,row+1):
    values = [ws.cell(row=i,column=j).value for j in range(2,column+1)]
    for j in range(len(values)):
        if (j+2) in skip:
            sdss_score+=values[j]
        else:
            score+=values[j]
            if (j+2) in pes:
                pes_score+=values[j]
            elif (j+2) in po:
                po_score+=values[j]
            else:
                print("There is an error for ", j)
            
                
    ws.cell(row=i, column=column+2, value = score)
    ws.cell(row=i, column=column+3, value = pes_score)
    ws.cell(row=i, column=column+4, value = po_score)
    ws.cell(row=i, column=column+5, value = sdss_score)
    print("i ",i, "     ",score,pes_score,po_score,pes_score+po_score)
    score = 0
    pes_score = 0
    po_score = 0
    sdss_score = 0

print("done")
wb.save('Respsheet.xlsx')#file name
    
