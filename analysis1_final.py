# -*- coding: utf-8 -*-
"""
Created on Sat Apr 26 09:00:46 2025
1) For each indicator/dimension: estimate the
percentage of children deprived (category deprevation) headcount and percentange

2)Multiple indicators under the same dimension
are aggregated in one dimension using union,
intersection, or intermediate cut-off
approaches.( Domain Deprevation) headcount and percentage

3)Estimate the percentage of children deprived in a number of dimensions 
( For each child how many dimensions are they deprived in) headcount and percentage

4)How many children are deprived in 0, 1, 2 or more dimensions?

-------------------------------------------------------------            
To Do: Nothing                                              -                                            
-------------------------------------------------------------

note: can't merge analysis 2 and 3 because the loops are in different orders


@author: krati


    
"""
import openpyxl
import functions as fn
import pandas as pd


"""
variables
"""
 # generate a list of numbers from 1 to the number of categories
d_n = range(1,11) #generate a list of numbers from 1 to the number of domain
ch_n = range(1,14) #generate the list containting the index of the children
cat_num = []
dom_num = list(d_n)
child_num = list(ch_n)
num0 = 0
num1 = 0
num2 = 0
num3 = 0
"""
dictionaries to store headcount
"""
cat_dep = {} # dictionary to store the list of the children deprived in each category 
dom_dep = {} # dictionary to store the list of the children deprived in each domain
ch_dom_dep={} #For each child how many dimensions are they deprived in
no_dom ={} #How many children are deprived in 0, 1, 2 or more dimensions
"""
dictionary to store percentages
"""
per_cat_dep ={}
per_dom_dep ={}
per_ch_dom_dep ={}
per_no_dom={}

"""
counts
"""

count_cat_dep = {}
count_dom_dep = {}
count_ch_dom_dep = {}

"""
lists that store temp data
"""
listoflist = [] 
perlist = []
count_list = []
"""
Workbook and Sheets
"""
wb = openpyxl.load_workbook('mafan.xlsx') #write your file name

wscat = wb['male_cat'] #score sheet
wsdom = wb['male_dom'] #domain score


"""
loop for anlysis 1 calculating the number and percentage of children deprived in each category
Done 
"""
yesno = [7,9,13,14,15,16,17,18,19,20,21,22,23,24,25,26,28,29,30]
oftsome = [10,11]
ai = [31]
tapwel = [27]
food = [12]
max_row = wscat.max_row #number of children
max_col = wscat.max_column #number of questions
cutoff = 0
for j in range(7,max_col+1):
    lis = [] #list for storing the children who are deprived in each category
    if j == 8: # skip the income question
        continue
    else:
        cat_num.append(j)
        for i in range(2,max_row+1):
            
            val = float(wscat.cell(row=i, column=j).value)
            if j in yesno:
                cutoff = fn.vul_cut_cat(val,"Yes")
            elif j in oftsome:
                cutoff = fn.vul_cut_cat(val,"Often")
            elif j in tapwel:
                cutoff = fn.vul_cut_cat(val,"Tap")
            elif j in food:
                cutoff = fn.vul_cut_cat(val,"food")
            elif j in ai:
                cutoff = fn.vul_cut_cat(val,"31")
            else:
                print("there is an error")
                cutoff = 1000
                
            if cutoff == 1:
                lis.append(i-1)
        per = (len(lis)/len(child_num))*100 
        perlist.append(per)
        count_list.append(len(lis))
        listoflist.append(lis)
for key, val in zip(cat_num, listoflist): # store the data in the dictionary
    cat_dep[key] = val
for key,val in zip(cat_num,perlist): # store the percentage data in the dictionary
    per_cat_dep[key] = val
for key,val in zip(cat_num,count_list):
    count_cat_dep[key]=val
count_list.clear()
perlist.clear()
listoflist.clear()

"""
percat_dep_df = pd.DataFrame(data=per_cat_dep,index=[0])
percat_dep_df = (percat_dep_df.T)
countcat_dep_df = pd.DataFrame(data=count_cat_dep,index=[0])
countcat_dep_df = (countcat_dep_df.T)


with pd.ExcelWriter('female_loop1.xlsx') as writer:
    percat_dep_df.to_excel(writer,sheet_name='female_cat_a2')
    countcat_dep_df.to_excel(writer,sheet_name='female_cat_a3')

"""
"""
loop for anlysis 2 calculating the number and percentage of children deprived in each domain
Done
"""

max_row = wsdom.max_row #number of children
max_col = wsdom.max_column #number of domains  
for j in range(2,max_col +1):
    lis = [] #list for storing the children who are deprived in each domain
    #print(wsdom.cell(row=1, column=j))
    for i in range(2,max_row +1):   
        val =int(wsdom.cell(row=i, column=j).value)
        #print(val,i,j)
        if val == 1:
                lis.append(i-1)
    per = (len(lis)/len(child_num))*100 
    count_list.append(len(lis))
    perlist.append(per)
    listoflist.append(lis)
for key, val in zip(dom_num, listoflist): # store the data in the dictionary
    dom_dep[key] = val
for key,val in zip(dom_num,perlist): # store the percentage data in the dictionary
    per_dom_dep[key] = val
for key,val in zip(dom_num,count_list):
    count_dom_dep[key]=val
count_list.clear()
perlist.clear()
listoflist.clear() 

"""
perdom_dep_df = pd.DataFrame(data=per_dom_dep,index=[0])
perdom_dep_df = (perdom_dep_df.T)
countdom_dep_df = pd.DataFrame(data=count_dom_dep,index=[0])
countdom_dep_df = (countdom_dep_df.T)


with pd.ExcelWriter('female_loop2.xlsx') as writer:
    #cat_dep_df.to_excel(writer,sheet_name='male_cat_a')
    perdom_dep_df.to_excel(writer,sheet_name='female_dom_a2')
    countdom_dep_df.to_excel(writer,sheet_name='female_dom_a3')
print(dom_dep,"         ",per_dom_dep,"            ",count_dom_dep)
"""

"""
loop for anlysis 3 For each child how many dimensions are they deprived in) headcount and percentage
Done

"""

for i in range(2,max_row+1):
    lis = [] #list for storing the dimension in which the child is deprived
    for j in range(2,max_col+1):

        val = int(wsdom.cell(row=i, column=j).value)
        if val == 1:
            lis.append(j-1)
    per = (len(lis)/len(dom_num))*100 
    perlist.append(per)
    count_list.append(len(lis))
    listoflist.append(lis)
for key, val in zip(child_num, listoflist): # store the data in the dictionary
    ch_dom_dep[key] = val
for key,val in zip(child_num,perlist): # store the percentage data in the dictionary
    per_ch_dom_dep[key] = val
for key,val in zip(child_num,count_list):
    count_ch_dom_dep[key]=val
count_list.clear()
perlist.clear()
listoflist.clear() 
"""
perdom_dep_df = pd.DataFrame(data=per_ch_dom_dep,index=[0])
perdom_dep_df = (perdom_dep_df.T)
countdom_dep_df = pd.DataFrame(data=count_ch_dom_dep,index=[0])
countdom_dep_df = (countdom_dep_df.T)


with pd.ExcelWriter('female_loop3.xlsx') as writer:
    perdom_dep_df.to_excel(writer,sheet_name='female_dom_a2')
    countdom_dep_df.to_excel(writer,sheet_name='female_dom_a3')
#print(ch_dom_dep,"         ",per_ch_dom_dep,"             ",count_ch_dom_dep)
"""
"""
loop for anlysis 4 How many children are deprived in 0, 1, 2 or more dimensions?
Done
"""



for i in range(1,len(child_num)+1):
    val = len(ch_dom_dep.get(i))
    if(val==0):
       num0+=1
    elif(val==1):
        num1+=1
    elif(val==2):
        num2+=1
    else:
        num3+=1
        
no_dom.update({0:num0})
no_dom.update({1:num1})
no_dom.update({2:num2})
no_dom.update({"More":num3})

#print(no_dom)
print(cat_dep,"              ")
print(dom_dep, "                   ", )
print(ch_dom_dep,"                 ")
