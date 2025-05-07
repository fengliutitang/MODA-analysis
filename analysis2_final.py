# -*- coding: utf-8 -*-
"""
Created on Tue Apr 29 10:44:45 2025

@author: krati
"""

import upsetplot
import itertools
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

wb = openpyxl.load_workbook('Respsheet.xlsx') #write your file name

wsdom = wb['dom_f']

"""
variables
"""
categories = []
for i in range(1,10):
    categories.append(str(i))
dom_dep = {} # dictionary to store the list of the children deprived in each domain
listoflist = [] 

"""
loop for finding out how many what are the number and indices of kids vulnerable in each domain
"""
max_row = wsdom.max_row +1 #number of children
max_col = wsdom.max_column+1 #number of domains  
for j in range(2,max_col):
    lis = [] #list for storing the children who are deprived in each domain
    for i in range(2,max_row):   
        val =int(wsdom.cell(row=i, column=j).value)
        if val == 1:
                lis.append(i-1)
    listoflist.append(lis)
for key, val in zip(categories, listoflist): # store the data in the dictionary
    dom_dep[key] = val


comb_list_list = []
comb_intersection_length_list = []
# identify per category combination the intersection length
for i in range(len(categories)):
    comb_list = list(itertools.combinations(categories, i+1))
    for elem in comb_list:
        comb_list_list.append(elem)
        # create a list of lists of categories for which to search the intersection length
        cat_lists = [dom_dep[x] for x in elem]
        comb_intersection_length_list.append(len(set(cat_lists[0]).intersection(*cat_lists)))



# remove category combinations with 0 intersections.
comb_list_list = np.asarray(comb_list_list,dtype="object")
comb_intersection_length_list = np.array(comb_intersection_length_list)
comb_list_list = comb_list_list[comb_intersection_length_list != 0]
comb_intersection_length_list = comb_intersection_length_list[comb_intersection_length_list != 0]



# create a membership data series which indicates the intersection size between the different sets
mem_series = upsetplot.from_memberships(comb_list_list,
                                        data=comb_intersection_length_list)

upsetplot.plot(mem_series,
               orientation='horizontal',
               show_counts=True,min_subset_size=9.0)

plt.savefig("f_full.pdf")
