# -*- coding: utf-8 -*-
"""
Created on Sat Apr 26 18:09:16 2025
Affect of gender
seperate male and female data into seperate sheets

@author: krati
"""
import openpyxl 
wb = openpyxl.load_workbook('Respsheet.xlsx') #file name here
wsread = wb['CtS'] # sheet with score data
wsfem = wb['Vf_an'] #sheet to store female data
wsmale = wb['Vm_an'] #sheet to store male data
max_row = wsread.max_row #number of children
max_col = wsread.max_column #number of categories
f_row = 2
m_row = 2
flag = 0
values = [wsread.cell(row=1,column=i).value for i in range(1,max_col+1)]

"""
headings
"""
for k in range(len(values)):
    wsfem.cell(row = 1, column = k+1).value = values[k]
    wsmale.cell(row = 1, column = k+1).value = values[k]
"""
loop for male and female
"""

for i in range(2,max_row+1):
             for j in range(1,max_col+1):
                    # copy the data in new sheet as a number
                    if(wsread.cell(row=i,column=4).value==1):
                        val = wsread.cell(row=i,column=j).value 
                        wsfem.cell(row=f_row, column=j, value = val)
                    
                    else:
                        val = wsread.cell(row=i,column=j).value 
                        wsmale.cell(row=m_row, column=j, value = val)
                        
                        flag = 1
                        
             if(flag==0):
                 f_row+=1
             else:
                 m_row+=1
                 flag=0
                       
print("done")
wb.save('Respsheet.xlsx')  