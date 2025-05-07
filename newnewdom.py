# -*- coding: utf-8 -*-
"""
Created on Wed Apr 30 18:50:15 2025

@author: krati
"""

import openpyxl
import functions as fn


dom_ques = [[7],[10,9,11],[13,29],[23,24,25],[11,12],[11,14,15,28],[25,26],[27,28],[16,17,18,19,20,21,22],[30,31]]
dom_head = ["Statelessness","Financial Security","Home","Personal Safetly","Diet","Health and Hygine","Work Exploitation","Water ","Education","Access to Information"]

dom_dic = {}
wb = openpyxl.load_workbook('mafan.xlsx') #write your file name

wsdom = wb['male_dom'] # sheet name where domain data is stored
ws = wb['male_cat'] # sheet name where score data is saved

"""
code to give heading in the domain worksheet
"""

wsdom.cell(row=1,column=1,value = "Child Id")


for i in range(len(dom_head)):
    wsdom.cell(row=1,column=i+2,value = dom_head[i])

"""
making dictionary
"""
for i in range(len(dom_head)):    
    dom_dic[dom_head[i]] = dom_ques[i]

max_row = ws.max_row #number of children
score_list = []
for k in range(2,max_row+1):
    stri = "Child"+str(k-1)
    wsdom.cell(row=k,column=1,value = stri)
    for i in range(len(dom_head)):
        domain = dom_head[i]
        val = dom_dic.get(domain)
        #print("row",k,"domain",i+1,"categories",val)
        for j in range(len(val)):
            score_list.append(ws.cell(row=k, column=val[j]).value) 
        #print("category score",score_list)
        dom_score = fn.vul_domain(score_list, i+1)
        wsdom.cell(row=k, column=i+2, value=dom_score)
        score_list.clear()
                  
        
print("done")
wb.save('mafan.xlsx') # write file name